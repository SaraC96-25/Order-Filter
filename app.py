import json
import re
from collections import defaultdict
from datetime import datetime, time
from io import BytesIO
from zoneinfo import ZoneInfo
from pathlib import Path

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


API_VERSION = "2026-04"
BASE_DIR = Path(__file__).resolve().parent
LOGO_PNG_PATH = BASE_DIR / "assets" / "wowstampa_logo.png"
BRAND_GREEN = "3AAA35"
BRAND_DARK = "1D1D1B"
LIGHT_GREEN = "EAF6E9"
LIGHT_GREY = "F3F5F7"


def get_secret(name: str, default: str = "") -> str:
    try:
        return st.secrets.get(name, default)
    except Exception:
        return default


SHOPIFY_SHOP_DOMAIN = get_secret("SHOPIFY_SHOP_DOMAIN")
SHOPIFY_ACCESS_TOKEN = get_secret("SHOPIFY_ACCESS_TOKEN")


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value):
    return normalize_text(value).casefold().replace(":", "").replace("_", " ").strip()


def gql_request(query: str, variables: dict | None = None) -> dict:
    if not SHOPIFY_SHOP_DOMAIN or not SHOPIFY_ACCESS_TOKEN:
        raise RuntimeError(
            "Configura SHOPIFY_SHOP_DOMAIN e SHOPIFY_ACCESS_TOKEN nei secrets di Streamlit."
        )

    domain = SHOPIFY_SHOP_DOMAIN.replace("https://", "").replace("http://", "").strip("/")
    url = f"https://{domain}/admin/api/{API_VERSION}/graphql.json"

    response = requests.post(
        url,
        headers={
            "Content-Type": "application/json",
            "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        },
        json={"query": query, "variables": variables or {}},
        timeout=60,
    )

    if response.status_code != 200:
        raise RuntimeError(f"Errore HTTP Shopify {response.status_code}: {response.text}")

    payload = response.json()
    if payload.get("errors"):
        raise RuntimeError(json.dumps(payload["errors"], indent=2, ensure_ascii=False))

    return payload["data"]


ORDERS_QUERY = """
query OrdersForReport($first: Int!, $after: String, $query: String!) {
  orders(first: $first, after: $after, query: $query, sortKey: CREATED_AT) {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
        name
        createdAt
        cancelledAt
        displayFinancialStatus
        displayFulfillmentStatus
        lineItems(first: 250) {
          pageInfo {
            hasNextPage
            endCursor
          }
          edges {
            node {
              id
              name
              title
              quantity
              currentQuantity
              sku
              variantTitle
              customAttributes {
                key
                value
              }            }
          }
        }
      }
    }
  }
}
"""


ORDER_LINE_ITEMS_QUERY = """
query OrderLineItems($orderId: ID!, $first: Int!, $after: String) {
  node(id: $orderId) {
    ... on Order {
      lineItems(first: $first, after: $after) {
        pageInfo {
          hasNextPage
          endCursor
        }
        edges {
          node {
            id
            name
            title
            quantity
            currentQuantity
            sku
            variantTitle
            customAttributes {
              key
              value
            }          }
        }
      }
    }
  }
}
"""


def fetch_more_line_items(order_id: str, after: str) -> list[dict]:
    all_items = []
    cursor = after

    while cursor:
        data = gql_request(
            ORDER_LINE_ITEMS_QUERY,
            {"orderId": order_id, "first": 250, "after": cursor},
        )
        line_items = data["node"]["lineItems"]
        all_items.extend([edge["node"] for edge in line_items["edges"]])
        if not line_items["pageInfo"]["hasNextPage"]:
            break
        cursor = line_items["pageInfo"]["endCursor"]

    return all_items


def fetch_orders(shopify_query: str) -> list[dict]:
    orders = []
    cursor = None

    progress = st.progress(0, text="Lettura ordini da Shopify...")
    pages = 0

    while True:
        data = gql_request(
            ORDERS_QUERY,
            {"first": 100, "after": cursor, "query": shopify_query},
        )

        conn = data["orders"]
        for edge in conn["edges"]:
            order = edge["node"]
            line_items_conn = order["lineItems"]
            line_items = [li_edge["node"] for li_edge in line_items_conn["edges"]]

            if line_items_conn["pageInfo"]["hasNextPage"]:
                extra_items = fetch_more_line_items(
                    order["id"], line_items_conn["pageInfo"]["endCursor"]
                )
                line_items.extend(extra_items)

            order["lineItems_flat"] = line_items
            orders.append(order)

        pages += 1
        progress.progress(
            min(95, pages * 10),
            text=f"Lettura ordini da Shopify... pagine lette: {pages}, ordini: {len(orders)}",
        )

        if not conn["pageInfo"]["hasNextPage"]:
            break
        cursor = conn["pageInfo"]["endCursor"]

    progress.progress(100, text=f"Completato: {len(orders)} ordini letti.")
    return orders


def build_shopify_date_query(start_date, end_date, timezone_name: str, exclude_cancelled: bool) -> str:
    tz = ZoneInfo(timezone_name)
    start_local = datetime.combine(start_date, time.min).replace(tzinfo=tz)
    end_local = datetime.combine(end_date, time.max).replace(tzinfo=tz)

    start_utc = start_local.astimezone(ZoneInfo("UTC")).isoformat().replace("+00:00", "Z")
    end_utc = end_local.astimezone(ZoneInfo("UTC")).isoformat().replace("+00:00", "Z")

    parts = [f"created_at:>={start_utc}", f"created_at:<={end_utc}"]
    if exclude_cancelled:
        parts.append("-status:cancelled")
    return " ".join(parts)


def product_title_from_line_item(item: dict) -> str:
    # Versione senza read_products:
    # usa solo i campi del line item disponibili con read_orders.
    return normalize_text(item.get("title")) or normalize_text(item.get("name"))


def product_matches(product_title: str, wanted_products: list[str], mode: str) -> bool:
    if not wanted_products:
        return True

    product_cf = product_title.casefold()
    wanted_cf = [p.casefold() for p in wanted_products if p.strip()]

    if mode == "Esatto":
        return product_cf in wanted_cf

    return any(w in product_cf for w in wanted_cf)



def extract_units_per_item(product_title: str, line_item_name: str = "") -> tuple[int, str]:
    """
    Estrae il moltiplicatore dal nome prodotto/riga ordine.

    Esempi:
    - "10 T-shirt Economy" -> 10
    - "20 T-shirt Economy" -> 20
    - "100 Shopper Cotone" -> 100

    Se non trova un numero iniziale, ritorna 1.
    """
    candidates = [normalize_text(product_title), normalize_text(line_item_name)]

    for candidate in candidates:
        if not candidate:
            continue

        match = re.match(r"^\s*(\d+)\s*(?:x|×|-)?\s+", candidate, flags=re.IGNORECASE)
        if match:
            value = int(match.group(1))
            if value > 0:
                return value, "numero iniziale nel nome prodotto"

    return 1, "fallback 1"


def parse_attribute_json(value: str) -> dict:
    value = normalize_text(value)
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except Exception:
        return {}

    if isinstance(parsed, dict):
        return parsed

    return {}


def extract_color(item: dict, color_keys: list[str]) -> tuple[str, str]:
    """
    Restituisce (colore, sorgente).

    Versione senza read_products:
    1. cerca in customAttributes / line item properties, tipico per opzioni personalizzate VOPO;
    2. usa variantTitle come fallback, se contiene pattern tipo "Colore: Nero".
    """

    normalized_color_keys = {normalize_key(k) for k in color_keys if normalize_key(k)}

    # 1) Custom attributes / line item properties
    for attr in item.get("customAttributes") or []:
        key = normalize_text(attr.get("key"))
        value = normalize_text(attr.get("value"))

        if normalize_key(key) in normalized_color_keys and value:
            return value, f"customAttributes.{key}"

        # Alcune app salvano un JSON in una property.
        nested = parse_attribute_json(value)
        for nested_key, nested_value in nested.items():
            if normalize_key(nested_key) in normalized_color_keys and normalize_text(nested_value):
                return normalize_text(nested_value), f"customAttributes.{key}.{nested_key}"

    # 2) Fallback su variantTitle: "Colore: Nero", "Taglia M / Nero", ecc.
    variant_title = normalize_text(item.get("variantTitle"))
    if variant_title and variant_title.casefold() != "default title":
        for color_key in color_keys:
            pattern = rf"{re.escape(color_key)}\s*[:=-]\s*([^,/|]+)"
            match = re.search(pattern, variant_title, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip(), "variantTitle pattern"

    return "", ""


def build_report(
    orders: list[dict],
    wanted_products: list[str],
    product_match_mode: str,
    color_keys: list[str],
    quantity_field: str,
    multiply_by_pack_size: bool,
    include_without_color: bool,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    summary = defaultdict(int)
    detail_rows = []

    for order in orders:
        for item in order.get("lineItems_flat", []):
            product_title = product_title_from_line_item(item)
            if not product_matches(product_title, wanted_products, product_match_mode):
                continue

            color, color_source = extract_color(item, color_keys)
            if not color and not include_without_color:
                continue

            if not color:
                color = "Senza colore"

            ordered_line_quantity = int(item.get(quantity_field) or 0)
            units_per_item, units_source = extract_units_per_item(
                product_title=product_title,
                line_item_name=item.get("name"),
            )

            counted_quantity = (
                ordered_line_quantity * units_per_item
                if multiply_by_pack_size
                else ordered_line_quantity
            )

            key = color
            summary[key] += counted_quantity

            detail_rows.append(
                {
                    "Ordine": order.get("name"),
                    "Data ordine": order.get("createdAt"),
                    "Prodotto": product_title,
                    "SKU": item.get("sku"),
                    "Nome riga ordine": item.get("name"),
                    "Quantità righe ordine": ordered_line_quantity,
                    "Pezzi per prodotto": units_per_item if multiply_by_pack_size else 1,
                    "Sorgente moltiplicatore": units_source if multiply_by_pack_size else "disattivato",
                    "Quantità totale calcolata": counted_quantity,
                    "Colore": color,
                    "Sorgente colore": color_source,
                    "Financial status": order.get("displayFinancialStatus"),
                    "Fulfillment status": order.get("displayFulfillmentStatus"),
                }
            )

    summary_rows = [
        {"Colore": color, "Quantità totale": qty}
        for color, qty in summary.items()
    ]

    summary_df = pd.DataFrame(summary_rows).sort_values(
        by=["Colore"], kind="stable"
    ) if summary_rows else pd.DataFrame(columns=["Colore", "Quantità totale"])

    detail_df = pd.DataFrame(detail_rows)
    return summary_df, detail_df


def dataframe_to_excel(summary_df: pd.DataFrame, detail_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Riepilogo")
        detail_df.to_excel(writer, index=False, sheet_name="Dettaglio ordini")

        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    return output.getvalue()



def _format_date_ita(d) -> str:
    return d.strftime("%d/%m/%Y")


def _write_dataframe_styled(ws, df: pd.DataFrame, start_row: int, start_col: int = 1):
    if df.empty:
        ws.cell(row=start_row, column=start_col, value="Nessun dato disponibile")
        ws.cell(row=start_row, column=start_col).font = Font(italic=True, color="666666")
        return start_row + 2

    thin = Side(style="thin", color="D9E1E5")
    header_fill = PatternFill("solid", fgColor=BRAND_GREEN)
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor=LIGHT_GREY)

    columns = list(df.columns)

    for j, col_name in enumerate(columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        is_alt = (i - (start_row + 1)) % 2 == 1
        for j, value in enumerate(row, start=start_col):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_alt:
                cell.fill = alt_fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, col in enumerate(columns, start=start_col):
        values = df.iloc[:, idx - start_col].fillna("").astype(str).tolist()
        max_len = max([len(str(col))] + [len(v) for v in values])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 3, 14), 46)

    return start_row + len(df) + 2


def create_branded_excel(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    start_date,
    end_date,
    selected_products: list[str],
    quantity_field: str,
    multiply_by_pack_size: bool,
    include_without_color: bool,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.sheet_view.showGridLines = False
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    if LOGO_PNG_PATH.exists():
        img = XLImage(str(LOGO_PNG_PATH))
        img.width = 190
        img.height = 89
        ws.add_image(img, "A1")

    ws.merge_cells("C2:G2")
    ws["C2"] = "REPORT COLORI ORDINI"
    ws["C2"].font = Font(size=20, bold=True, color=BRAND_DARK)
    ws["C2"].alignment = Alignment(horizontal="right")

    ws.merge_cells("C3:G3")
    ws["C3"] = f"Periodo analizzato: {_format_date_ita(start_date)} – {_format_date_ita(end_date)}"
    ws["C3"].font = Font(size=11, bold=True, color=BRAND_GREEN)
    ws["C3"].alignment = Alignment(horizontal="right")

    ws.merge_cells("C4:G4")
    ws["C4"] = f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
    ws["C4"].font = Font(size=10, color="666666")
    ws["C4"].alignment = Alignment(horizontal="right")

    box_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    for row in range(7, 11):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = box_fill

    ws.merge_cells("A7:G7")
    ws["A7"] = "PARAMETRI REPORT"
    ws["A7"].font = Font(bold=True, color=BRAND_DARK)

    products_text = ", ".join(selected_products) if selected_products else "Tutti i prodotti filtrati per data"

    ws.merge_cells("A8:D8")
    ws["A8"] = f"Prodotti inclusi: {products_text}"
    ws["A8"].alignment = Alignment(wrap_text=True)

    ws.merge_cells("E8:G8")
    ws["E8"] = f"Campo quantità: {quantity_field}"

    ws.merge_cells("A9:D9")
    ws["A9"] = f"Moltiplicatore da nome prodotto: {'Sì' if multiply_by_pack_size else 'No'}"

    ws.merge_cells("E9:G9")
    ws["E9"] = f"Righe senza colore incluse: {'Sì' if include_without_color else 'No'}"

    total_qty = int(summary_df["Quantità totale"].sum()) if not summary_df.empty else 0
    total_colors = int(summary_df["Colore"].nunique()) if not summary_df.empty else 0

    ws["A11"] = "Totale pezzi"
    ws["A11"].font = Font(bold=True, color=BRAND_DARK)
    ws["B11"] = total_qty
    ws["B11"].font = Font(size=14, bold=True, color=BRAND_GREEN)

    ws["D11"] = "Colori distinti"
    ws["D11"].font = Font(bold=True, color=BRAND_DARK)
    ws["E11"] = total_colors
    ws["E11"].font = Font(size=14, bold=True, color=BRAND_GREEN)

    ws["A13"] = "RIEPILOGO UNIFICATO PER COLORE"
    ws["A13"].font = Font(size=14, bold=True, color=BRAND_DARK)
    next_row = _write_dataframe_styled(ws, summary_df, start_row=14, start_col=1)

    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=7)
    note_cell = ws.cell(row=next_row, column=1)
    note_cell.value = "Nota: il riepilogo somma tutti i prodotti filtrati e aggrega i totali esclusivamente per colore."
    note_cell.font = Font(italic=True, color="666666")
    note_cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A14"

    ws2 = wb.create_sheet("Dettaglio ordini")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = "DETTAGLIO ORDINI"
    ws2["A1"].font = Font(size=16, bold=True, color=BRAND_DARK)
    ws2["A2"] = f"Periodo: {_format_date_ita(start_date)} – {_format_date_ita(end_date)}"
    ws2["A2"].font = Font(size=10, color=BRAND_GREEN)
    _write_dataframe_styled(ws2, detail_df, start_row=4, start_col=1)
    ws2.freeze_panes = "A5"

    ws3 = wb.create_sheet("Riepilogo dati")
    ws3.sheet_view.showGridLines = False
    _write_dataframe_styled(ws3, summary_df, start_row=1, start_col=1)
    ws3.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_branded_pdf(
    summary_df: pd.DataFrame,
    start_date,
    end_date,
    selected_products: list[str],
    quantity_field: str,
    multiply_by_pack_size: bool,
    include_without_color: bool,
) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCustom",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1D1D1B"),
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    sub_style = ParagraphStyle(
        "SubCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#3AAA35"),
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    body_style = ParagraphStyle(
        "BodyCustom",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor("#1D1D1B"),
        alignment=TA_LEFT,
        leading=13,
    )

    elements = []

    if LOGO_PNG_PATH.exists():
        logo = RLImage(str(LOGO_PNG_PATH), width=5.2 * cm, height=2.45 * cm)
        logo.hAlign = "CENTER"
        elements.append(logo)

    elements.append(Paragraph("REPORT COLORI ORDINI", title_style))
    elements.append(Paragraph(f"Periodo analizzato: {_format_date_ita(start_date)} – {_format_date_ita(end_date)}", sub_style))

    products_text = ", ".join(selected_products) if selected_products else "Tutti i prodotti filtrati per data"
    info_text = (
        f"<b>Generato il:</b> {datetime.now().strftime('%d/%m/%Y alle %H:%M')}<br/>"
        f"<b>Prodotti inclusi:</b> {products_text}<br/>"
        f"<b>Campo quantità:</b> {quantity_field}<br/>"
        f"<b>Moltiplicatore da nome prodotto:</b> {'Sì' if multiply_by_pack_size else 'No'}<br/>"
        f"<b>Righe senza colore incluse:</b> {'Sì' if include_without_color else 'No'}"
    )
    elements.append(Paragraph(info_text, body_style))
    elements.append(Spacer(1, 0.35 * cm))

    total_qty = int(summary_df["Quantità totale"].sum()) if not summary_df.empty else 0
    total_colors = int(summary_df["Colore"].nunique()) if not summary_df.empty else 0

    kpi_table = Table(
        [["Totale pezzi", f"{total_qty}", "Colori distinti", f"{total_colors}"]],
        colWidths=[4.0 * cm, 3.0 * cm, 4.0 * cm, 3.0 * cm],
    )
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EAF6E9")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1D1D1B")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D9E1E5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#D9E1E5")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.45 * cm))

    section_style = ParagraphStyle(
        "SectionCustom",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.HexColor("#1D1D1B"),
        spaceAfter=8,
    )
    elements.append(Paragraph("RIEPILOGO UNIFICATO PER COLORE", section_style))

    table_rows = [["Colore", "Quantità totale"]]
    if summary_df.empty:
        table_rows.append(["Nessun dato disponibile", ""])
    else:
        for _, row in summary_df.iterrows():
            table_rows.append([str(row["Colore"]), int(row["Quantità totale"])])

    summary_table = Table(table_rows, colWidths=[10 * cm, 4 * cm])
    summary_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3AAA35")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#D9E1E5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#D9E1E5")),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]

    for row_i in range(1, len(table_rows)):
        if row_i % 2 == 0:
            summary_style.append(("BACKGROUND", (0, row_i), (-1, row_i), colors.HexColor("#F3F5F7")))

    summary_table.setStyle(TableStyle(summary_style))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.35 * cm))

    note_style = ParagraphStyle(
        "NoteCustom",
        parent=styles["Italic"],
        fontName="Helvetica-Oblique",
        fontSize=8.5,
        textColor=colors.HexColor("#666666"),
    )
    elements.append(Paragraph(
        "Nota: il riepilogo somma tutti i prodotti filtrati e aggrega i totali esclusivamente per colore.",
        note_style,
    ))

    doc.build(elements)
    return buffer.getvalue()


def dataframe_to_excel(summary_df: pd.DataFrame, detail_df: pd.DataFrame) -> bytes:
    # Compatibilità con eventuali chiamate vecchie.
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Riepilogo")
        detail_df.to_excel(writer, index=False, sheet_name="Dettaglio ordini")
    return output.getvalue()


st.set_page_config(page_title="Report Shopify colori", page_icon="🎨", layout="wide")

st.title("🎨 Report Shopify: quantità ordinate per colore")
st.caption(
    "Filtra gli ordini Shopify per data e prodotto, calcola i pezzi reali e unifica il risultato finale per colore, senza distinguere per prodotto."
)

with st.sidebar:
    st.header("Impostazioni")

    timezone_name = st.text_input("Timezone negozio", value="Europe/Rome")

    st.subheader("Date ordine")
    start_date = st.date_input("Da")
    end_date = st.date_input("A")

    exclude_cancelled = st.checkbox("Escludi ordini cancellati", value=True)

    st.subheader("Prodotti")
    product_text = st.text_area(
        "Titoli prodotto da includere, uno per riga",
        placeholder="20 T-shirt Economy\nFelpa Premium",
    )
    product_match_mode = st.radio(
        "Match prodotto",
        ["Contiene", "Esatto"],
        index=0,
        horizontal=True,
    )

    st.subheader("Campo colore")
    color_keys_text = st.text_input(
        "Nomi possibili del campo colore",
        value="Colore, Color, colour, colore",
        help="Se VOPO salva il campo con un nome diverso, aggiungilo qui.",
    )

    quantity_label = st.radio(
        "Quantità ordini da conteggiare",
        ["currentQuantity (al netto di rimborsi/rimozioni)", "quantity (quantità originale)"],
        index=0,
    )

    multiply_by_pack_size = st.checkbox(
        "Moltiplica per il numero iniziale nel nome prodotto",
        value=True,
        help="Esempio: se il prodotto è '10 T-shirt' e la quantità ordinata è 3, il totale diventa 10 × 3 = 30.",
    )

    include_without_color = st.checkbox("Mostra anche righe senza colore", value=False)

    run = st.button("Genera report", type="primary")


wanted_products = [line.strip() for line in product_text.splitlines() if line.strip()]
color_keys = [x.strip() for x in color_keys_text.split(",") if x.strip()]
quantity_field = "currentQuantity" if quantity_label.startswith("currentQuantity") else "quantity"


with st.expander("Come impostare i secrets su Streamlit Cloud"):
    st.code(
        """
SHOPIFY_SHOP_DOMAIN = "tuo-negozio.myshopify.com"
SHOPIFY_ACCESS_TOKEN = "shpat_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
        """.strip(),
        language="toml",
    )
    st.write("Non mettere mai il token dentro GitHub. Usa i secrets di Streamlit Cloud.")


if run:
    if start_date > end_date:
        st.error("La data iniziale non può essere successiva alla data finale.")
        st.stop()

    if not wanted_products:
        st.warning("Non hai indicato prodotti: il report includerà tutti i prodotti trovati negli ordini.")

    try:
        shopify_query = build_shopify_date_query(
            start_date=start_date,
            end_date=end_date,
            timezone_name=timezone_name,
            exclude_cancelled=exclude_cancelled,
        )

        st.info(f"Query Shopify: `{shopify_query}`")

        orders = fetch_orders(shopify_query)

        summary_df, detail_df = build_report(
            orders=orders,
            wanted_products=wanted_products,
            product_match_mode=product_match_mode,
            color_keys=color_keys,
            quantity_field=quantity_field,
            multiply_by_pack_size=multiply_by_pack_size,
            include_without_color=include_without_color,
        )

        st.subheader("Riepilogo unificato per colore")
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

        st.subheader("Dettaglio ordini")
        st.dataframe(detail_df, use_container_width=True, hide_index=True)

        st.info("I file di export includono un layout istituzionale Wowstampa con logo, periodo analizzato, parametri del report e riepilogo finale per colore.")

        branded_excel_bytes = create_branded_excel(
            summary_df=summary_df,
            detail_df=detail_df,
            start_date=start_date,
            end_date=end_date,
            selected_products=wanted_products,
            quantity_field=quantity_field,
            multiply_by_pack_size=multiply_by_pack_size,
            include_without_color=include_without_color,
        )
        branded_pdf_bytes = create_branded_pdf(
            summary_df=summary_df,
            start_date=start_date,
            end_date=end_date,
            selected_products=wanted_products,
            quantity_field=quantity_field,
            multiply_by_pack_size=multiply_by_pack_size,
            include_without_color=include_without_color,
        )

        col_dl1, col_dl2, col_dl3 = st.columns(3)
        with col_dl1:
            st.download_button(
                "Scarica Excel istituzionale",
                data=branded_excel_bytes,
                file_name=f"wowstampa_report_colori_{start_date}_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_dl2:
            st.download_button(
                "Scarica PDF istituzionale",
                data=branded_pdf_bytes,
                file_name=f"wowstampa_report_colori_{start_date}_{end_date}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        with col_dl3:
            csv_bytes = summary_df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                "Scarica CSV riepilogo",
                data=csv_bytes,
                file_name=f"report_shopify_colori_{start_date}_{end_date}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with st.expander("Debug: campi colore trovati nei line item"):
            debug_rows = []
            for order in orders[:25]:
                for item in order.get("lineItems_flat", []):
                    product_title = product_title_from_line_item(item)
                    if not product_matches(product_title, wanted_products, product_match_mode):
                        continue
                    debug_rows.append(
                        {
                            "Ordine": order.get("name"),
                            "Prodotto": product_title,
                            "Nome riga ordine": item.get("name"),
                            "Variant title": item.get("variantTitle"),
                            "Selected options": "Non lette: manca scope read_products",
                            "Custom attributes": json.dumps(
                                item.get("customAttributes") or [],
                                ensure_ascii=False,
                            ),
                        }
                    )
            st.dataframe(pd.DataFrame(debug_rows), use_container_width=True, hide_index=True)

    except Exception as exc:
        st.error(str(exc))
        st.stop()
